import re
import unicodedata
from datetime import datetime
from typing import Tuple, List, Dict, Any

from bs4 import BeautifulSoup


# =========================
# Utilidades
# =========================
def to_snake(s: str) -> str:
    s = (s or "").strip()
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s)

    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.lower()

    s = s.replace("°", "")
    s = s.replace("/", " ")
    s = s.replace("-", " ")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace(" ", "_")

    return s or "col"


def make_unique(headers: List[str]) -> List[str]:
    out = []
    seen = {}
    for h in headers:
        k = to_snake(h)
        if k in seen:
            seen[k] += 1
            k = f"{k}_{seen[k]}"
        else:
            seen[k] = 1
        out.append(k)
    return out


def try_parse_date(s: str):
    s = (s or "").strip()
    if not s:
        return None

    # cortar hora si viene "2025-09-01 00:00:00"
    if " " in s:
        s2 = s.split(" ")[0].strip()
        # si parece fecha, úsala
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s2) or re.fullmatch(r"\d{2}/\d{2}/\d{4}", s2) or re.fullmatch(r"\d{2}-\d{2}-\d{4}", s2):
            s = s2

    fmts = [
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
    ]

    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except:
            pass

    return None



def headers_have_pair_pattern(headers: List[str]) -> bool:
    """True si headers vienen en pares iguales: A,A,B,B,C,C ..."""
    if not headers or len(headers) % 2 != 0:
        return False
    for i in range(0, len(headers), 2):
        if headers[i] != headers[i+1]:
            return False
    return True


def collapse_row_by_pairs(row: List[str]) -> List[str]:
    """Toma una celda de cada par: 0,2,4..."""
    return [row[i] for i in range(0, len(row), 2)]

def _is_blank(v: str) -> bool:
    s = (v or "").strip().lower()
    return s == "" or s in {"none", "nan", "null", "n/a", "na"}

def maybe_collapse_pairs_by_empty_columns(headers: List[str], rows: List[List[str]], empty_ratio: float = 0.90):
    """
    Colapsa pares (col1/col2, col3/col4, ...) cuando la columna derecha está casi siempre vacía.
    """
    if not headers or not rows:
        return headers, rows
    if len(headers) % 2 != 0:
        return headers, rows

    n = len(headers)
    total = len(rows)
    if total == 0:
        return headers, rows

    right_empty = [0] * (n // 2)

    for r in rows:
        if len(r) < n:
            r = r + [""] * (n - len(r))
        for i in range(0, n, 2):
            right = r[i+1]
            if _is_blank(str(right)):
                right_empty[i//2] += 1

    ratios = [x / total for x in right_empty]

    # colapsar si al menos 70% de pares tiene derecha vacía >= empty_ratio
    if sum(1 for rr in ratios if rr >= empty_ratio) >= int(0.7 * (n // 2)):
        headers2 = [headers[i] for i in range(0, n, 2)]
        rows2 = []
        for r in rows:
            if len(r) < n:
                r = r + [""] * (n - len(r))
            rows2.append([r[i] for i in range(0, n, 2)])
        return headers2, rows2

    return headers, rows



#========================
# colapsar columnas para xls
#=======================

def drop_blank_duplicate_columns(headers: List[str], rows: List[List[str]], blank_ratio: float = 0.95):
    """
    Elimina columnas cuyo header termina en _2, _3, etc. y que están vacías casi siempre.
    Pensado para XLS donde el colapso por pares falla por columnas fantasma.
    """
    if not headers or not rows:
        return headers, rows

    n = len(headers)
    total = len(rows)
    if total == 0:
        return headers, rows

    def is_blank(v):
        s = "" if v is None else str(v)
        s = s.replace("\xa0", " ").replace("\u200b", " ").strip().lower()
        return s == "" or s in {"nan", "none", "null", "nat"}

    # calcular ratio de vacíos por columna
    empty_counts = [0] * n
    for r in rows:
        rr = list(r)
        if len(rr) < n:
            rr += [""] * (n - len(rr))
        elif len(rr) > n:
            rr = rr[:n]

        for i, v in enumerate(rr):
            if is_blank(v):
                empty_counts[i] += 1

    # decidir columnas a conservar
    keep_idx = []
    for i, h in enumerate(headers):
        # detectar headers duplicados generados por make_unique: algo_2, algo_3...
        is_dup = bool(re.search(r"_\d+$", h))
        empty_ratio_col = empty_counts[i] / total

        # si es duplicada y casi siempre vacía -> eliminar
        if is_dup and empty_ratio_col >= blank_ratio:
            continue

        keep_idx.append(i)

    # reconstruir
    headers2 = [headers[i] for i in keep_idx]
    rows2 = []
    for r in rows:
        rr = list(r)
        if len(rr) < n:
            rr += [""] * (n - len(rr))
        elif len(rr) > n:
            rr = rr[:n]
        rows2.append([rr[i] for i in keep_idx])

    return headers2, rows2




def collapse_pairs_xls_force(headers: List[str], rows: List[List[str]], empty_ratio: float = 0.70):
    if not headers or not rows or len(headers) % 2 != 0:
        return headers, rows

    n = len(headers)
    total = len(rows)
    if total == 0:
        return headers, rows

    def norm_cell(v):
        s = "" if v is None else str(v)
        return s.replace("\xa0", " ").replace("\u200b", " ").strip()

    def is_blank(v):
        s = norm_cell(v).lower()
        return s == "" or s in {"nan", "none", "null", "nat"}

    # Normaliza todas las filas al tamaño n
    norm_rows = []
    for r in rows:
        rr = [norm_cell(x) for x in (r or [])]
        if len(rr) < n:
            rr = rr + [""] * (n - len(rr))
        elif len(rr) > n:
            rr = rr[:n]
        norm_rows.append(rr)

    pairs = n // 2
    good_pairs = 0

    # Evalúa por par si la columna derecha es vacía casi siempre
    for i in range(0, n, 2):
        empty_count = 0
        for rr in norm_rows:
            if is_blank(rr[i+1]):
                empty_count += 1
        if (empty_count / total) >= empty_ratio:
            good_pairs += 1

    # Si al menos 50% de pares cumplen, colapsa
    if good_pairs >= int(0.6 * pairs):
        headers2 = [headers[i] for i in range(0, n, 2)]
        rows2 = [[rr[i] for i in range(0, n, 2)] for rr in norm_rows]
        return headers2, rows2

    return headers, norm_rows





def normalize_identificacion(headers_norm: List[str], rows: List[List[str]]) -> None:
    """
    Corrige cédulas que llegaron como número (sin 0 inicial) a string de 10 dígitos.
    Solo se aplica si existe tipo_identificacion y/o número_identificacion.
    """
    # posibles nombres normalizados
    tipo_keys = ["tipo_identificacion", "tipo_de_identificacion"]
    num_keys = ["numero_identificacion", "numero_de_identificacion"]

    tipo_idx = next((headers_norm.index(k) for k in tipo_keys if k in headers_norm), None)
    num_idx = next((headers_norm.index(k) for k in num_keys if k in headers_norm), None)

    if num_idx is None:
        return

    for r in rows:
        if num_idx >= len(r):
            continue

        num = (r[num_idx] or "").strip()
        if not num:
            continue

        # limpiar .0 (cuando Excel lo pasa como float)
        if re.fullmatch(r"\d+\.0", num):
            num = num.split(".")[0]

        # si es todo dígitos
        if not re.fullmatch(r"\d+", num):
            r[num_idx] = num
            continue

        # determinar si es cédula por tipo_identificacion si existe
        is_cedula = False
        if tipo_idx is not None and tipo_idx < len(r):
            t = (r[tipo_idx] or "").strip().upper()
            if t in ["CEDULA", "CÉDULA", "CI", "C.I."]:
                is_cedula = True

        # si no hay tipo, pero tiene 9 dígitos -> asumimos cédula ecuatoriana y rellenamos
        if is_cedula and len(num) < 10:
            r[num_idx] = num.zfill(10)
        elif (not is_cedula) and len(num) == 9:
            r[num_idx] = num.zfill(10)
        else:
            r[num_idx] = num


# =========================
# Parse HTML PRASS (estable)
# =========================
def clean_header_text(th) -> str:
    txt = th.get_text(" ", strip=True)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def expand_headers_html(ths):
    expanded = []
    original = []
    for th in ths:
        txt = clean_header_text(th)
        original.append(txt)
        colspan = th.get("colspan")
        try:
            c = int(colspan) if colspan else 1
        except:
            c = 1
        for _ in range(max(1, c)):
            expanded.append(txt)
    return original, expanded


def expand_row_cells_html(tr):
    cells = tr.find_all(["td", "th"])
    out = []
    for c in cells:
        txt = c.get_text(" ", strip=True)
        txt = re.sub(r"\s+", " ", txt).strip()
        colspan = c.get("colspan")
        try:
            k = int(colspan) if colspan else 1
        except:
            k = 1
        for _ in range(max(1, k)):
            out.append(txt)
    return out


def parse_prass_html(html_bytes: bytes) -> Dict[str, Any]:
    """
    IMPORTANTE: aquí NO modificamos headers por fila.
    Detectamos si headers vienen duplicados (pares) y colapsamos UNA sola vez,
    luego aplicamos el mismo colapso a todas las filas.
    """
    soup = BeautifulSoup(html_bytes, "html.parser")
    table = soup.find("table")
    if not table:
        raise ValueError("No se encontró <table> en el HTML.")

    thead = table.find("thead")
    tbody = table.find("tbody")
    if not thead or not tbody:
        raise ValueError("La tabla no tiene thead/tbody.")

    header_tr = thead.find("tr")
    if not header_tr:
        raise ValueError("No se encontró fila de encabezados en thead.")

    ths = header_tr.find_all("th")
    headers_original, headers_exp = expand_headers_html(ths)

    # Detectar patrón por pares y colapsar headers UNA sola vez
    pair_mode = headers_have_pair_pattern(headers_exp)
    headers_final = collapse_row_by_pairs(headers_exp) if pair_mode else headers_exp

    rows_final = []
    for tr in tbody.find_all("tr"):
        row_exp = expand_row_cells_html(tr)

        # Si estamos en modo pares, intentamos colapsar fila si corresponde
        if pair_mode and len(row_exp) >= 2 and len(row_exp) % 2 == 0:
            row_use = collapse_row_by_pairs(row_exp)
        else:
            row_use = row_exp

        rows_final.append(row_use)

    headers_norm = make_unique(headers_final)

    # normalizar longitudes
    col_count = len(headers_norm)
    fixed_rows = []
    for r in rows_final:
        rr = list(r)[:col_count] + [""] * max(0, col_count - len(r))
        fixed_rows.append(rr)

    # corregir cédulas si aplica
    normalize_identificacion(headers_norm, fixed_rows)

    return {
        "tipo_detectado": "HTML_PRASS",
        "headers_originales": headers_original,
        "headers_normalizados": headers_norm,
        "rows": fixed_rows
    }


# =========================
# Parse XLSX (2 filas header + merges)
# =========================
def bytes_to_file(content: bytes):
    import io
    return io.BytesIO(content)


def parse_prass_xlsx(content: bytes) -> Dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=bytes_to_file(content), data_only=True)
    ws = wb.active

    def get_row_values(row_idx: int) -> List[str]:
        vals = []
        max_col = ws.max_column
        for c in range(1, max_col + 1):
            v = ws.cell(row=row_idx, column=c).value
            vals.append("" if v is None else str(v).strip())

        # fill-left para merges
        for i in range(1, len(vals)):
            if vals[i] == "" and vals[i-1] != "":
                vals[i] = vals[i-1]
        return vals

    h1 = get_row_values(1)
    h2 = get_row_values(2)

    headers_raw = []
    for a, b in zip(h1, h2):
        a = (a or "").strip()
        b = (b or "").strip()
        if a and b and b.lower() not in a.lower():
            headers_raw.append(f"{a} {b}")
        else:
            headers_raw.append(a or b or "")

    # Colapsar si vienen duplicados por pares
    pair_mode = headers_have_pair_pattern(headers_raw)
    headers_final = collapse_row_by_pairs(headers_raw) if pair_mode else headers_raw

    headers_norm = make_unique(headers_final)

    # datos desde fila 3
    rows = []
    max_col = len(headers_final)

    for r in range(3, ws.max_row + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            vals.append("" if v is None else str(v).strip())

        # colapsar fila si aplica
        if pair_mode and len(vals) % 2 == 0:
            vals = collapse_row_by_pairs(vals)

        vals = vals[:max_col] + [""] * max(0, max_col - len(vals))

        if any(x.strip() for x in vals):
            rows.append(vals)

    # colapsar pares por columnas derechas vacías (si aplica)
    headers_final, rows = maybe_collapse_pairs_by_empty_columns(headers_final, rows)

    # normalizar headers (por si cambió)
    headers_norm = make_unique(headers_final)

    # corregir identificación (ya con índices finales)
    normalize_identificacion(headers_norm, rows)    
        
    return {
        "tipo_detectado": "XLSX",
        "headers_originales": headers_final,
        "headers_normalizados": headers_norm,
        "rows": rows
    }


# =========================
# Parse XLS binario (2 filas header)
# =========================
def parse_prass_xls(content: bytes) -> Dict[str, Any]:
    import io
    import pandas as pd

    df = pd.read_excel(io.BytesIO(content), header=None, dtype=str, engine="xlrd").fillna("")
    if df.shape[0] < 2:
        raise ValueError("El XLS no tiene suficientes filas para encabezados.")

    row1 = [str(x).strip() for x in df.iloc[0].tolist()]
    row2 = [str(x).strip() for x in df.iloc[1].tolist()]

    # fill-left
    for i in range(1, len(row1)):
        if row1[i] == "" and row1[i-1] != "":
            row1[i] = row1[i-1]
    for i in range(1, len(row2)):
        if row2[i] == "" and row2[i-1] != "":
            row2[i] = row2[i-1]

    headers_raw = []
    for a, b in zip(row1, row2):
        a = (a or "").strip()
        b = (b or "").strip()
        if a and b and b.lower() not in a.lower():
            headers_raw.append(f"{a} {b}")
        else:
            headers_raw.append(a or b or "")

    pair_mode = headers_have_pair_pattern(headers_raw)
    headers_final = collapse_row_by_pairs(headers_raw) if pair_mode else headers_raw
    headers_norm = make_unique(headers_final)

    # datos desde fila 3
    data_df = df.iloc[2:].copy()
    rows = []
    max_col = len(headers_final)

    for _, r in data_df.iterrows():
        vals = [str(x).strip() for x in r.tolist()]

        if pair_mode and len(vals) % 2 == 0:
            vals = collapse_row_by_pairs(vals)

        vals = vals[:max_col] + [""] * max(0, max_col - len(vals))
        if any(x.strip() for x in vals):
            rows.append(vals)

    # colapsar pares forzado si la derecha está vacía
    headers_final, rows = collapse_pairs_xls_force(headers_final, rows)

    # regenerar headers_norm tras colapso
    headers_norm = make_unique(headers_final)


    # corregir identificación (ya con índices finales)
    normalize_identificacion(headers_norm, rows)

    # eliminar columnas duplicadas (_2, _3...) que están vacías
    headers_norm, rows = drop_blank_duplicate_columns(headers_norm, rows, blank_ratio=0.95)



    return {
        "tipo_detectado": "XLS",
        "headers_originales": headers_final,
        "headers_normalizados": headers_norm,
        "rows": rows
    }


# =========================
# Detector principal
# =========================
def parse_any_prass_file(filename: str, content: bytes) -> Dict[str, Any]:
    name_l = (filename or "").lower()
    head = content[:4096]
    head_text = head.decode("utf-8", errors="ignore").lower()
    looks_like_html = ("<table" in head_text) or ("<html" in head_text) or ("<!doctype html" in head_text)

    if looks_like_html:
        return parse_prass_html(content)

    if name_l.endswith(".xlsx"):
        return parse_prass_xlsx(content)

    if name_l.endswith(".xls"):
        return parse_prass_xls(content)

    # fallback
    try:
        return parse_prass_xlsx(content)
    except:
        return parse_prass_xls(content)


# =========================
# Stats (fecha_vacunacion)
# =========================
def extract_fecha_vacunacion_stats(headers_norm: List[str], rows: List[List[str]]):
    # aceptar variantes del header
    candidates = [
        "fecha_vacunacion",
        "fecha_vacunacion_vacunador",
        "fecha_vacunacion_vacunacion",
    ]

    idx = None
    for c in candidates:
        if c in headers_norm:
            idx = headers_norm.index(c)
            break

    # fallback: cualquier header que empiece con fecha_vacunacion
    if idx is None:
        for i, h in enumerate(headers_norm):
            if h.startswith("fecha_vacunacion"):
                idx = i
                break

    if idx is None:
        return None, None, []

    dates = []
    for r in rows:
        s = r[idx] if idx < len(r) else ""
        d = try_parse_date(s)
        if d:
            dates.append(d)

    if not dates:
        return None, None, []

    min_d = min(dates)
    max_d = max(dates)
    months = sorted({f"{d.year:04d}-{d.month:02d}" for d in dates})
    return min_d, max_d, months
